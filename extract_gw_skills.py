
import openpyxl
import json
import os
import re

file_path = r'd:\王炜\工作\个人工作\wx_proj\new_zx_calc\doc\伤害计算 - 鬼王 - 北辰 - 副本.xlsx'

def extract_info():
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    try:
        # Load the excel file directly using openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        all_data = {}

        # Map Chinese labels to our keys
        mapping = {
            "附加固定攻击": "SkillAttackFixedBonus",
            "附加固定真气": "SkillManaFixedBonus",
            "附加固定气血": "SkillHealthFixedBonus",
            "附加固定防御": "SkillDefenseFixedBonus",
            "附加固定攻击比": "SkillAttackPercentBonus",
            "附加固定真气比": "SkillManaPercentBonus",
            "附加固定气血比": "SkillHealthPercentBonus",
            "附加固定防御比": "SkillDefensePercentBonus",
            "附加爆伤": "SkillCriticalDamagePercentBonus"
        }

        # Specific sheets we want
        target_sheets = ['未名禅伤害计算', '九变伤害计算']

        for sheet_name in target_sheets:
            if sheet_name not in wb.sheetnames:
                continue
                
            ws = wb[sheet_name]
            
            info = {
                "SkillName": sheet_name,
                "SkillAttackFixedBonus": 0,
                "SkillManaFixedBonus": 0,
                "SkillHealthFixedBonus": 0,
                "SkillDefenseFixedBonus": 0,
                "SkillAttackPercentBonus": 0,
                "SkillManaPercentBonus": 0,
                "SkillHealthPercentBonus": 0,
                "SkillDefensePercentBonus": 0,
                "SkillCriticalDamagePercentBonus": 0
            }

            # Search in the sheet
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str):
                        # Clean label for comparison
                        clean_val = val.replace('\n', '').replace(' ', '')
                        for label, key in mapping.items():
                            if label in clean_val:
                                try:
                                    # Case 1: Value is in the same cell
                                    # Matches both integer and float, and potentially percentages
                                    matches = re.findall(r'(\d+\.?\d*)', clean_val)
                                    if matches:
                                        target_val = float(matches[0])
                                    else:
                                        # Case 2: Value is in the next cell
                                        next_cell_col = cell.column + 1
                                        target_val = ws.cell(row=cell.row, column=next_cell_col).value
                                        
                                        if target_val is not None:
                                            if isinstance(target_val, (int, float)):
                                                target_val = float(target_val)
                                            elif isinstance(target_val, str):
                                                m = re.findall(r'(\d+\.?\d*)', target_val)
                                                if m:
                                                    target_val = float(m[0])
                                                else:
                                                    target_val = 0
                                            else:
                                                target_val = 0
                                        else:
                                            target_val = 0
                                    
                                    # Only update if we found a value and it's not already set
                                    # (sometimes multiple matches occur, we want the most specific one)
                                    if target_val is not None and target_val > 0:
                                        # Heuristic for percentages: if value is between 0 and 5, it might be a multiplier (e.g. 2.2 instead of 220%)
                                        # but let's just extract what's there and let the user decide.
                                        info[key] = target_val
                                except Exception:
                                    pass

            all_data[sheet_name] = info

        print(json.dumps(all_data, indent=2, ensure_ascii=False))

    except Exception as e:
        print(f"Error processing excel: {e}")

if __name__ == "__main__":
    extract_info()
