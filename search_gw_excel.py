
import openpyxl
import os

file_path = r'd:\王炜\工作\个人工作\wx_proj\new_zx_calc\doc\伤害计算 - 鬼王 - 新20250730.xlsx'

def search_excel():
    if not os.path.exists(file_path):
        print(f"Error: File not found at {file_path}")
        return

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"--- Sheet: {sheet_name} ---")
            found = False
            for row in ws.iter_rows():
                for cell in row:
                    val = cell.value
                    if val and isinstance(val, str):
                        if any(keyword in val for keyword in ['未名', '鬼王', '九变']):
                            print(f"  Found '{val}' at cell {cell.coordinate}")
                            found = True
            if not found:
                print("  (No relevant keywords found)")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    search_excel()
